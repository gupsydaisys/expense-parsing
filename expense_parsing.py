from openpyxl import load_workbook

from sets import Set

#######################################################################################################################
################################################### STEP 1 !!!!!!!  ####################################################
# 1. Look for TODO
#######################################################################################################################

# TODO CONFIGURE: Set to name of sheet that references other sheets, must be the same directory as this program
INDEX_SHEET_NAME = 'Index.xlsx'
INDEX_WB = load_workbook(INDEX_SHEET_NAME)

# TODO CONFIGURE: Set to name of Sheet holding a week worth of expenses
E0 = 'Expenses0.xlsx'
E1 = 'Expenses1.xlsx'
E0_WB = load_workbook(E0)
E1_WB = load_workbook(E1)

# TODO Configure: If you added a workbook, add it this array as well
EXPENSE_WBS = [E0_WB, E1_WB]

# A set of every item name that is entered in as an expense
all_item_names = Set()

# @return a list of all unique item names across all sheets
def get_all_unique_item_names():
  for wb in EXPENSE_WBS:
    for sheet_name in wb.sheetnames:
      add_item_names(wb[sheet_name])
  return all_item_names

def add_item_names(sheet):
  for col_row in sheet['B']:
    if col_row.value is None:
      continue
    all_item_names.add(col_row.value.strip())
  return True

#######################################################################################################################
################################################### STEP 2 !!!!!!!  ####################################################
# 1. lOOK FOR TODO
# Note: The easiest way to see items not mapped to is by running `map_label_to_row_on_sheet()` and looking at the errors (line 103)
#######################################################################################################################

# This is a list of all labels that we use to label a type of expense.  Note that the actual label may not be in this
# set however every actual label should map to something in this set.
all_possible_item_labels = Set()

# @return a map of each item name to a label
def map_item_names_to_label():
  label_map = {}
  label_map_sheet = INDEX_WB['LabelMap']
  for index in range(1, 500):
    if label_map_sheet[index][0].value is None:
      continue
    item_name = label_map_sheet[index][0].value.strip()
    label = label_map_sheet[index][1].value
    label_map[item_name] = label
    all_possible_item_labels.add(label)
  return label_map

# Map of every time name to a label
LABEL_MAP = map_item_names_to_label()

# TODO: Uncomment out the lines below to see what new mappings you need to add to LabelMap
# all_item_names = set(get_all_unique_item_names())
# all_mapped_item_names = set(LABEL_MAP.values() + LABEL_MAP.keys())
# new_item_names_to_add_to_map = all_item_names - all_mapped_item_names

#######################################################################################################################
################################################### STEP 3 !!!!!!!  ####################################################
# 1. Look for TODO
#######################################################################################################################

# @return a map where the keys are a label and value are an array of row data
# eg:
# grocery: [
#            {
#              name: 'haight street market',
#              cost: '130',
#              sheet: 'meow cat meow sheet name',
#            },
#            ...
#          ]
def map_label_to_row_on_sheet():
  cray_zay_map = {}
  for wb in EXPENSE_WBS:
    for sheet_name in wb.sheetnames:
      sheet = wb[sheet_name]
      last_index = 100
      for index in range(1, last_index):
        item_name = sheet[index][1].value
        if item_name is None:
          continue
        key_label = ''
        if item_name.strip() in LABEL_MAP:
          key_label = LABEL_MAP[item_name.strip()]
        elif item_name.strip() in all_possible_item_labels:
          key_label = item_name.strip()
        else:
          # print item_name
          print "Error!!!! The following item was not mapped to a label: " + item_name
          # print "Make sure you add that item_name to LabelMaps in Google Sheets"
        cost = sheet[index][2].value
        if (key_label not in cray_zay_map):
          cray_zay_map[key_label] = []
        cray_zay_map[key_label].append({'name': item_name, 'cost': cost, 'sheet': sheet_name })
  return cray_zay_map

# TODO uncomment out the below line
# MAP_OF_LABELS_TO_ITEM_ROWS = map_label_to_row_on_sheet()

def calculate_costs_per_label():
  mappy = {}
  for key in MAP_OF_LABELS_TO_ITEM_ROWS.keys():
    yolo = sum(MAP_OF_LABELS_TO_ITEM_ROWS[key][i]['cost'] for i in range(0, len(MAP_OF_LABELS_TO_ITEM_ROWS[key])))
    mappy[key] = yolo
  return mappy

# TODO uncomment out the below line
# MEOW = calculate_costs_per_label()

print "TODO: Check for negatives and go through ? and categorize"

def cost_highest_to_lowest():
  return sorted(MEOW, key=MEOW.get, reverse=True)

# TODO: Uncomment these out get printed a list of items highest spend to lowest spend
# and then a list of the cost associated with each

# costs_h_to_l = cost_highest_to_lowest()
# for x in costs_h_to_l:
  # print x

# for x in costs_h_to_l:
  # print MEOW[x]
